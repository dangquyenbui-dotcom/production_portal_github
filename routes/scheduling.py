# routes/scheduling.py
"""
Production Scheduling routes
Handles display and updates for the production scheduling grid.
ADDED: API endpoint to export detailed FG inventory for summary cards.
ADDED: API endpoint to export detailed Shipped Orders for summary card.
"""

from flask import Blueprint, render_template, jsonify, request, session, redirect, url_for, flash, send_file
from auth import require_login, require_scheduling_admin, require_scheduling_user
from routes.main import validate_session
# UPDATED IMPORT: Added ERP service getter
from database import scheduling_db, get_erp_service
import traceback
import openpyxl
from io import BytesIO
from datetime import datetime, timedelta # Added timedelta

# The url_prefix makes this blueprint's routes available under '/scheduling'
scheduling_bp = Blueprint('scheduling', __name__, url_prefix='/scheduling')
erp_service = get_erp_service() # Get ERP service instance

@scheduling_bp.route('/')
@validate_session
def index():
    # ... (existing code remains the same) ...
    """Renders the main production scheduling grid page."""
    if not require_login(session):
        return redirect(url_for('main.login'))

    if not (require_scheduling_admin(session) or require_scheduling_user(session)):
        flash('Scheduling privileges are required to access this module.', 'error')
        return redirect(url_for('main.dashboard'))

    # Fetch data from ERP joined with local projections
    data = scheduling_db.get_schedule_data()

    # Unpack the dictionary to pass its contents as separate variables to the template
    return render_template(
        'scheduling/index.html',
        user=session['user'],
        schedule_data=data.get('grid_data', []),
        fg_on_hand_split=data.get('fg_on_hand_split', {}),
        shipped_current_month=data.get('shipped_current_month', 0),
        now=datetime.now()
    )


@scheduling_bp.route('/api/update-projection', methods=['POST'])
@validate_session
def update_projection():
    # ... (existing code remains the same) ...
    """API endpoint to save projection data from the grid."""
    if not require_scheduling_admin(session):
        return jsonify({'success': False, 'message': 'Edit permission required'}), 403

    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Invalid data received'}), 400

        # Extract data from the JSON payload
        so_number = data.get('so_number')
        part_number = data.get('part_number')
        risk_type = data.get('risk_type')
        quantity = data.get('quantity')
        username = session.get('user', {}).get('username', 'unknown')

        # Basic validation
        if not all([so_number, part_number, risk_type]):
             return jsonify({'success': False, 'message': 'Missing required fields: so_number, part_number, or risk_type'}), 400

        try:
            quantity = float(quantity) if quantity is not None else 0.0
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'Quantity must be a valid number'}), 400

        # Call the database method to perform the upsert
        success, message = scheduling_db.update_projection(
            so_number=so_number,
            part_number=part_number,
            risk_type=risk_type,
            quantity=quantity,
            username=username
        )

        if success:
            return jsonify({'success': True, 'message': message})
        else:
            return jsonify({'success': False, 'message': message}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'An internal server error occurred.'}), 500

@scheduling_bp.route('/api/export-xlsx', methods=['POST'])
@validate_session
def export_xlsx():
    # ... (existing code remains the same) ...
    """API endpoint to export the visible grid data to an XLSX file."""
    if not (require_scheduling_admin(session) or require_scheduling_user(session)):
        return jsonify({'success': False, 'message': 'Authentication required'}), 401

    try:
        data = request.get_json()
        headers = data.get('headers', [])
        rows = data.get('rows', [])

        if not headers or not rows:
            return jsonify({'success': False, 'message': 'No data to export'}), 400

        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Schedule Export"

        # Write headers
        ws.append(headers)

        # Write data rows, attempting to convert to numbers
        for row_data in rows:
            processed_row = []
            for cell_value in row_data:
                # If the value is a string, try to clean and convert it to a number
                if isinstance(cell_value, str):
                    cleaned_value = cell_value.replace('$', '').replace(',', '')
                    try:
                        # Try converting to float for decimals
                        processed_row.append(float(cleaned_value))
                    except (ValueError, TypeError):
                        # If it fails, it's not a number, so append the original string
                        processed_row.append(cell_value)
                else:
                    # If it's already a number (or None), append it as is
                    processed_row.append(cell_value)

            ws.append(processed_row)

        # Save the workbook to a BytesIO object
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"schedule_export_{timestamp}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'An error occurred during export.'}), 500


@scheduling_bp.route('/api/export-fg-details')
@validate_session
def export_fg_details():
    # ... (existing code remains the same) ...
    """API endpoint to export detailed FG inventory based on date buckets."""
    if not (require_scheduling_admin(session) or require_scheduling_user(session)):
        flash('Permission denied.', 'error')
        return redirect(url_for('main.dashboard')) # Redirect if accessed directly without permission

    bucket = request.args.get('bucket')
    if bucket not in ['prior', 'mid', 'recent']:
        flash('Invalid data bucket specified.', 'error')
        return redirect(url_for('.index')) # Redirect back to scheduling page

    try:
        # --- Date Calculation Logic (mirrors sales_queries.py) ---
        today = datetime.now()
        first_of_this_month = today.replace(day=1)
        last_of_previous_month = first_of_this_month - timedelta(days=1)
        prior_cutoff_date = last_of_previous_month.replace(day=21)
        current_cutoff_date = today.replace(day=21)

        # Use 'YYYY-MM-DD' format for query parameters
        sql_date_format = '%Y-%m-%d'
        prior_cutoff_str_sql = prior_cutoff_date.strftime(sql_date_format)
        current_cutoff_str_sql = current_cutoff_date.strftime(sql_date_format)

        start_date = None
        end_date = None
        filename_suffix = ""

        if bucket == 'prior':
            end_date = prior_cutoff_str_sql # Less than prior cutoff
            filename_suffix = "prior"
        elif bucket == 'mid':
            start_date = prior_cutoff_str_sql # Greater than or equal to prior cutoff
            end_date = current_cutoff_str_sql # Less than current cutoff
            filename_suffix = "mid"
        elif bucket == 'recent':
            start_date = current_cutoff_str_sql # Greater than or equal to current cutoff
            filename_suffix = "recent"

        # Fetch detailed inventory data from ERP
        inventory_data = erp_service.get_detailed_fg_inventory(start_date, end_date)

        if not inventory_data:
            flash(f'No inventory data found for the "{filename_suffix}" period.', 'info')
            return redirect(url_for('.index'))

        # --- Generate Excel ---
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"FG Inventory {filename_suffix.capitalize()}"

        # Define headers (using keys from the first row)
        headers = list(inventory_data[0].keys())
        ws.append(headers)

        # Write data rows
        for row_dict in inventory_data:
            row_values = [row_dict.get(h) for h in headers]
            ws.append(row_values)

        # Auto-size columns (optional, can be slow for large files)
        # for col in ws.columns:
        #     max_length = 0
        #     column = col[0].column_letter # Get the column name
        #     for cell in col:
        #         try: # Necessary to avoid error on empty cells
        #             if len(str(cell.value)) > max_length:
        #                 max_length = len(cell.value)
        #         except:
        #             pass
        #     adjusted_width = (max_length + 2)
        #     ws.column_dimensions[column].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Prepare filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"fg_inventory_detail_{filename_suffix}_{timestamp}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        traceback.print_exc()
        flash(f'An error occurred during the FG inventory export: {e}', 'error')
        return redirect(url_for('.index'))

# --- NEW ROUTE ---
@scheduling_bp.route('/api/export-shipped-details')
@validate_session
def export_shipped_details():
    """API endpoint to export detailed Shipped Orders for the current month."""
    if not (require_scheduling_admin(session) or require_scheduling_user(session)):
        flash('Permission denied.', 'error')
        return redirect(url_for('main.dashboard'))

    try:
        # Fetch detailed shipment data from ERP
        shipment_data = erp_service.get_detailed_shipments_current_month()

        if not shipment_data:
            flash(f'No shipment data found for the current month.', 'info')
            return redirect(url_for('.index'))

        # --- Generate Excel ---
        wb = openpyxl.Workbook()
        ws = wb.active
        month_name = datetime.now().strftime("%B_%Y")
        ws.title = f"Shipped_{month_name}"

        # Define headers (using keys from the first row)
        headers = list(shipment_data[0].keys())
        ws.append(headers)

        # Write data rows
        for row_dict in shipment_data:
            row_values = [row_dict.get(h) for h in headers]
            # Format numeric columns if needed (optional)
            for i, header in enumerate(headers):
                 if header in ['ShippedQuantity', 'UnitPrice', 'LineValue']:
                     try:
                         row_values[i] = float(row_values[i]) if row_values[i] is not None else None
                     except (ValueError, TypeError):
                         pass # Keep original value if conversion fails
            ws.append(row_values)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Prepare filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"shipped_details_{month_name}_{timestamp}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        traceback.print_exc()
        flash(f'An error occurred during the shipment export: {e}', 'error')
        return redirect(url_for('.index'))
# --- END NEW ROUTE ---